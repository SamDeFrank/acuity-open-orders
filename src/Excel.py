from Common import Common
from Orders import OrderReport, Order
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color

SAVE_PATH = r"C:/Users/Sam/Documents/Acuity Open Orders"

class ExcelFile(Common):

  MYSTERY_WIDTH_OFFSET = .71
  SHEET_NAME = "Orders"
  col_widths = {
    'A': 14.00 + MYSTERY_WIDTH_OFFSET,
    'B': 10.00 + MYSTERY_WIDTH_OFFSET,
    'C': 06.71 + MYSTERY_WIDTH_OFFSET,
    'D': 07.00 + MYSTERY_WIDTH_OFFSET,
    'E': 08.00 + MYSTERY_WIDTH_OFFSET,
    'F': 11.00 + MYSTERY_WIDTH_OFFSET,
  }

  styles = {
    'new': {
      'fill' : PatternFill(patternType="solid", fgColor=Color(rgb="FFFFFF00", tint=0.0, type='rgb'),bgColor=Color(rgb="00000000", tint=0.0, type='rgb')),
      'font' : Font()
    },
    'closed': {
      'fill' : PatternFill(),
      'font' : Font(strike=True) 
    },
    'recent': {
      'fill': PatternFill(patternType="solid", fgColor=Color(theme=0,tint=-0.249977111117893, type='theme'),bgColor=Color(indexed=64,tint=0.0,type='indexed')),
      'font': Font()
    },
    'open': {
      'fill' : PatternFill(),
      'font' : Font()
    },
    'header': {
      'font': Font(size=24, bold=True)
    }
  }




  @classmethod
  def make_new(cls, filename: str):
    """Generate new excel workbook
    :param filename: Name of file to be created"""

    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = cls.SHEET_NAME

    return cls(wb, filename)
  

  @classmethod
  def update_existing(cls, filename: str):
    """Load existing excel sheet to be updated
    :param file_to_update: name of excle file to load"""

    wb = load_workbook(filename)
    ws1 = wb[wb.sheetnames[0]]
    print(ws1['G1'].value.split(" "))
    ws1.title = "Delete Me"
    wb.create_sheet(cls.SHEET_NAME)


    # return cls(wb, filename, 1)


  def __init__(self, wb: Workbook, filename: str, date_created=None):
    self.wb = wb
    self.filename = filename
    ws = wb[self.SHEET_NAME]

    #format sheet
    ws.merge_cells("G1:I1")
    ws["G1"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    today_str = datetime.now().strftime('%m-%d-%y    %I:%M %p')

    if date_created:
      date_str = date_created.strftime('%m-%d-%y    %I:%M %p')
      created_str = f"Created:   {date_str}\nUpdated:   {today_str}"
    else:
      created_str = f"Created:   {today_str}"

    ws["G1"] = created_str
   
   
    for i in self.col_widths:
      ws.column_dimensions[i].width = self.col_widths[i]



  def write_order(self, ws, order: Order, row: int):
    for i, name in enumerate(self.col_names):

      cell = ws.cell(row=row, column=i+1)
      cell.value = order.get(name, row)
      cell.font = self.styles[order.status]['font']
      cell.fill = self.styles[order.status]['fill']
      cell.border = Border(bottom=Side(style='thin'))

      if name in ["Need-By Date", "Quantity Ordered", "Quantity Received", "Balance-Due"]:
        cell.alignment = Alignment(horizontal="right")


  def write_orders(self, all_orders: OrderReport):
    excel_row = 1;
    ws = self.wb[self.SHEET_NAME]
    for loc in self.ship_tos:

      orders = all_orders.by_ship_to(loc)

      if len(orders) == 0:
        continue

      cell = ws.cell(row=excel_row, column=1)
      cell.value = loc
      cell.font = self.styles['header']['font']
      excel_row += 1
      
      for order in orders:
        print(f"row value in write_orderes: {excel_row}")
        self.write_order(ws, order, excel_row)
        excel_row +=1
      
      excel_row += 1
    

  
  def save(self):
    self.wb.save(self.filename)










# def write(ws, orders, created_on, update, settings):
#   row_offset = 0
#   locations = sorted(orders.keys())


  

#   #the big loop
#   for location in locations:
#     sz = len(orders[location])

#     if(sz > 0):
#       row = 1 + row_offset
#       ws.row_dimensions[row].height = 30
#       ws['A%s' % row].font = style['fonts']['header']
#       ws['A%s' % row] = location

#       # insert 'created on' date
#       if row == 1:
#         date_string = "Created On:   {}".format(created_on.strftime('%m-%d-%Y'))
#         if update:
#           date_string += "\nUpdated On:   {}".format(datetime.date.today().strftime('%m-%d-%Y'))
#         ws["G1"] = date_string

#       row_offset += 1

#       for excelRow in range(1, sz+1):

#         #begin writing the business data to sheet   
#         for excelCol in range(1, len(COLUMN_NAMES) + 1):

#           col_name = COLUMN_NAMES[excelCol-1]
#           cell = ws.cell(row=excelRow + row_offset, column=excelCol)
#           value = orders[location][excelRow-1]['info'][col_name]
#           status = orders[location][excelRow-1]['status']

#           # right justify date
#           if col_name == "Need-By Date":
#             cell.alignment = Alignment(horizontal="right")

#           # generate 'balance due' formula
#           if col_name == "Balance Due":
#             value = "=if(C{row}-D{row}<0, 0, C{row}-D{row})".format(row=excelRow + row_offset)

#           if col_name in ["Quantity Ordered", "Quantity Received"]:
#             try:
#               value = int(value)
#             except:
#               pass

#           # add horizontal gridlines
#           cell.border = style['borders']['underline']

#           cell.font = style[status]['font']
#           cell.fill = style[status]['fill']
          
#           cell.value = value
          
#         # add gridlines to the right of the page for handwritten notes on print out
#         ws.cell(row=excelRow + row_offset, column=7).border = style['borders']['underline']
#         ws.cell(row=excelRow + row_offset, column=8).border = style['borders']['mid']
#         ws.cell(row=excelRow + row_offset, column=9).border = style['borders']['underline']

#       row_offset += sz + 1


orders = OrderReport.from_tsv(r"C:\Users\Sam\Downloads\export.tsv")

# sheet = ExcelFile.make_new("test.xlsx")
sheet = ExcelFile.update_existing(r"C:\Users\Sam\Documents\Acuity Open Orders\2020-11-14 Open Orders.xlsx")

# sheet.write_orders(orders)
# sheet.save()

