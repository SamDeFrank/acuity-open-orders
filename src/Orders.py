from Common import Common
from openpyxl import load_workbook
from datetime import datetime
# from openpyxl.styles import Color, PatternFill, Font, Border
# from openpyxl.styles.differential import DifferentialStyle
# from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule


class OrderReport(Common):

  yellow_fill = "FFFFFF00"
  grey_tint = -0.249977111117893

  def __init__(self, orders):

    self.orders = orders

  @classmethod
  def from_tsv(cls, path: str):
    """Load order data from a tsv located at {path}"""
    orders_dict = {}

    with open(path, 'r') as file:
      data = file.read()

    data = [row.split("\t") for row in data.split("\t\n")]
    data = [[feild.replace('"', '') for feild in row] for row in data]

    for row in data[1:]:
      if len(row) == len(data[0]):
        order = dict(zip(data[0], row))
        loc = cls.ship_to_mapping[order['Ship-To Location']]
        order.update({'Ship-To Location': loc, 'status': "open"})
        order = Order(order, cls.col_names)

        orders_dict.update({order.order_id(): order})

    return cls(orders_dict)

  @classmethod
  def from_xlsx(cls, file_path: str):
    """Load order data from an excel file at {path}"""
    location = ""
    orders_dict = {}
    
    wb = load_workbook(filename=file_path)
    ws = wb[wb.sheetnames[0]]

    for row in ws:
      status = "open"
      if row[0].value == None:
        continue
      if row[0].value in cls.ship_to_mapping:
        location = row[0].value
        continue
      values = [cell.value for cell in row]
      order = dict(zip(cls.col_names, values))

      if row[0].font.strike:
        status = "closed"
      elif row[0].fill.fgColor.rgb == cls.yellow_fill or row[0].fill.fgColor.tint == cls.grey_tint:
        status = "recent"

      order.update({'status': status, 'Ship-To Location': location})
      order = Order(order, cls.col_names)

      orders_dict.update({order.order_id(): order})

    return cls(orders_dict)
    

  def by_ship_to(self, location: str):
    output = []
    for _id in self.orders:
      order = self.orders[_id]
      if order.ship_to == location:
        output.append(order)
    return sorted(output, key=lambda x: x.sort_date())
  
  def update_from(self, new_orders):
    nords = new_orders.orders

    for oid in self.orders:
      if oid not in nords:
        self.orders[oid].status = "closed"
      else:
        status = self.orders[oid].status
        self.orders.update({oid :nords[oid]})
        self.orders[oid].status = status

    for noid in new_orders.orders:
      if noid not in self.orders:
        self.orders.update({noid: nords[noid]})
  
  def get_notes_from(self, path):
    pass
    #method should grab user notes from appropriate excel file




class Order():
  """Container class for storing and accessing Acuity purchase order data"""

  def __init__(self, order: dict, fields: list):

    self.bad_date = False
    self.status = order['status']
    self.ship_to = order['Ship-To Location']
    self.info = {i: order[i] if i in order else "" for i in fields}

    try:
      date = datetime.strptime(self.info['Need-By Date'],'%d-%b-%Y %H:%M:%S')
      self.info['Need-By Date'] = date
    except ValueError:
      self.bad_date = True


  def sort_date(self):
    if not self.bad_date:
      return self.info['Need-By Date']
    else:
      return datetime.strptime("01-jan-2000", '%d-%b-%Y')


  def order_id(self):
    return self.info['PO Number'] + self.info['Item Number']


  def get(self, attr:str, row=None):
    if attr == "Need-By Date":
      date = self.info[attr]
      return date if self.bad_date else date.strftime("%m-%d-%Y")
    elif attr == "Balance Due":
      return f"=if(C{row}-D{row}<0, 0, C{row}-D{row})"
    else:
      return self.info[attr]

 
  def display(self):
    max_len = len(max(self.info.keys(), key=len))
    banner = "="*(max_len - len('ORDER INFO')//2+1)
    print(f"{banner}ORDER INFO{banner}")
    for i in self.info:
      print(" "*(max_len-len(i)), i + ":", self.info[i])
    print("")
