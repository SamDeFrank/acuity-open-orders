import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color

TODAY = datetime.date.today()
OUTPUT_FILENAME = "\\%s Open Orders.xlsx" % TODAY
SHIP_TOs        = ['Fishers', 'Crawfordsville', 'Des Plaines', 'MPF', 'GPF', 'SEAC']
COLUMN_NAMES    = ["Item Number", "PO Number", "Quantity Ordered", "Quantity Received", "Balance Due", "Need-By Date", "G", "H", "I"]
MYSTERY_WIDTH_OFFSET = .71

style = {
  'fonts': {
    'header': Font(size=24, bold=True)
    },
  'borders': {
    'mid': Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')),
    'underline': Border(bottom=Side(style='thin'))
  },
  'new': {
    'fill' : PatternFill(patternType="solid", fgColor=Color(rgb="FFFFFF00", tint=0.0, type='rgb'),bgColor=Color(rgb="00000000", tint=0.0, type='rgb')),
    'font' : Font()
  },
  'closed': {
    'fill' : PatternFill(),
    'font' : Font(strike=True) 
  },
  'recent': {
    'fill': PatternFill(patternType="solid", fgColor=Color(theme=0,tint=-0.249977111117893, type='theme'),bgColor=Color(indexed=64,tint=0.0,type='indexed')),
    'font': Font()
  },
  'open': {
    'fill' : PatternFill(),
    'font' : Font()
  }
}


def valid_date(po):
  '''Used as 'key' when sorting purchase orders by due date'''    
  try:
    output = datetime.datetime.strptime(po['Need-By Date'].split()[0], '%d-%b-%Y')
  except ValueError:
    # if the acuity buyer forgets to put a due date on their PO, sort it as if its date were january 1st of 2000. but leave the cell blank
    output = datetime.datetime.strptime("01-jan-2000", '%d-%b-%Y')

  return output

def status(fill, font):
  """assignes status to order based on the current style of the cell"""
  if font.strike:
    return "closed"
  
  #new orders become recent and recent orders stay recent
  elif fill.fgColor.rgb == "FFFFFF00" or fill.fgColor.tint == -0.249977111117893:
    return "recent"
  
  else:
    return "open"

def compare(existing_orders, new_orders):
  updated_orders = existing_orders['orders'].copy()
  locations_to_sort = []

  #check for orders to update or close:
  for ship_to in updated_orders:
    for po in updated_orders[ship_to]:
      order_id = po['info']['id']
      if order_id not in new_orders['ids']:
        po['status'] = 'closed'
      else:
        info = {
          "Quantity Ordered": new_orders["ids"][order_id]["Quantity Ordered"],
          "Quantity Received": new_orders["ids"][order_id]["Quantity Received"],
          "Need-By Date": new_orders["ids"][order_id]["Need-By Date"]
        }
        po['info'].update(info)
  
  #check for new orders to add
  for ship_to in new_orders['orders']:
    for po in new_orders['orders'][ship_to]:
      if po['info']['id'] not in existing_orders['ids']:
        if ship_to not in locations_to_sort:
          locations_to_sort.append(ship_to)
        po['status'] = "new"
        updated_orders[ship_to].append(po)


  for loc in locations_to_sort:
    updated_orders[loc].sort(key=lambda x : datetime.datetime.strptime(x['info']['Need-By Date'].split()[0], '%m-%d-%Y'))
  
  return updated_orders

def transfer_user_notes(old_orders, new_orders):

  updated_orders = new_orders['orders'].copy()

  for ship_to in updated_orders:
    for po in updated_orders[ship_to]:
      order_id = po['info']['id']
      if order_id in old_orders['ids']:
        info = {
          'G': old_orders['ids'][order_id]['G'],
          'H': old_orders['ids'][order_id]['H'],
          'I': old_orders['ids'][order_id]['I']
        }
        po['info'].update(info)
  
  return updated_orders
        

def load_tsv(path):
  data_list = []
  orders = {i: [] for i in SHIP_TOs}
  order_ids = {}
  
  with open(path, 'r') as file:
    raw_data_list = [row.split("\t") for row in file.read().split("\t\n")] 
  
  for i in raw_data_list:
    data_list.append([k.replace('"', "") for k in i])
  
  column_headers = data_list[0]

  #parse data into list of dicts with column headers as the dict keys, then sort it by due date
  order_list = [dict(zip(column_headers, row)) for row in data_list[1:] if len(row) == len(column_headers)]
  order_list = sorted(order_list, key = valid_date)

  #group orders by ship-to-location in a dictionary where ship-to locations are the keys
  for po in order_list:
    
    #if the date string is not empty, reformat it so it's easier to read on the excel sheet   
    if len(po["Need-By Date"]) > 0:
      po["Need-By Date"] = datetime.datetime.strptime(po["Need-By Date"].split(" ")[0], '%d-%b-%Y').date().strftime('%m-%d-%Y')

    po["Balance Due"] = None

    #resolve ship-to location based on whats in the data. eg: both 'P1-CRAWFORDSVILLE' and 'P2-CRAWFORDSVILLE' are grouped under 'Crawfordsville'
    location = [loc for loc in SHIP_TOs if loc.upper() in po["Ship-To Location"].upper()][0]
    
    #trim unnecessary data from each order
    order = {k:v for (k,v) in po.items() if k in COLUMN_NAMES}
    order.update({"G": None, "H": None, "I": None, "id": order["PO Number"] + order["Item Number"]})

    orders[location].append({"info": order, "status": "open"})
    order_ids[order["id"]] = order


  return {"orders": orders, "ids": order_ids}

def load_xlsx(ws):
  
  location = ""
  orders = {i: [] for i in SHIP_TOs}
  order_ids = {}
  
  for row in ws:

    if row[0].value == None:
      continue

    if row[0].value in SHIP_TOs:
      location = row[0].value
      continue

    values = [cell.value for cell in row]
    order = dict(zip(COLUMN_NAMES, values))
    order.update({"id": order["PO Number"] + order["Item Number"]})

    orders[location].append({"info": order, "status": status(row[0].fill, row[0].font)})
    order_ids[order["id"]] = order
  
  return {"orders": orders, "ids": order_ids}

def write(ws, orders, created_on, update, settings):
  row_offset = 0
  locations = sorted(orders.keys())

  # set column widths
  ws.column_dimensions["A"].width = 14.00 + MYSTERY_WIDTH_OFFSET
  ws.column_dimensions["B"].width = 10.00 + MYSTERY_WIDTH_OFFSET
  ws.column_dimensions["C"].width = 06.71 + MYSTERY_WIDTH_OFFSET
  ws.column_dimensions["D"].width = 07.00 + MYSTERY_WIDTH_OFFSET
  ws.column_dimensions["E"].width = 08.00 + MYSTERY_WIDTH_OFFSET
  ws.column_dimensions["F"].width = 11.00 + MYSTERY_WIDTH_OFFSET

  ws.merge_cells("G1:I1")
  ws["G1"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

  #the big loop
  for location in locations:
    sz = len(orders[location])

    if(sz > 0):
      row = 1 + row_offset
      ws.row_dimensions[row].height = 30
      ws['A%s' % row].font = style['fonts']['header']
      ws['A%s' % row] = location

      # insert 'created on' date
      if row == 1:
        date_string = "Created On:   {}".format(created_on.strftime('%m-%d-%Y'))
        if update:
          date_string += "\nUpdated On:   {}".format(datetime.date.today().strftime('%m-%d-%Y'))
        ws["G1"] = date_string

      row_offset += 1

      for excelRow in range(1, sz+1):

        #begin writing the business data to sheet   
        for excelCol in range(1, len(COLUMN_NAMES) + 1):

          col_name = COLUMN_NAMES[excelCol-1]
          cell = ws.cell(row=excelRow + row_offset, column=excelCol)
          value = orders[location][excelRow-1]['info'][col_name]
          status = orders[location][excelRow-1]['status']

          # right justify date
          if col_name == "Need-By Date":
            cell.alignment = Alignment(horizontal="right")

          # generate 'balance due' formula
          if col_name == "Balance Due":
            value = "=if(C{row}-D{row}<0, 0, C{row}-D{row})".format(row=excelRow + row_offset)

          if col_name in ["Quantity Ordered", "Quantity Received"]:
            try:
              value = int(value)
            except:
              pass

          # add horizontal gridlines
          if settings['gridlines']:
            cell.border = style['borders']['underline']


          cell.font = style[status]['font']
          cell.fill = style[status]['fill']
          
          cell.value = value
          
        # add gridlines to the right of the page for handwritten notes on print out
        if settings['gridlines']:
          ws.cell(row=excelRow + row_offset, column=7).border = style['borders']['underline']
          ws.cell(row=excelRow + row_offset, column=8).border = style['borders']['mid']
          ws.cell(row=excelRow + row_offset, column=9).border = style['borders']['underline']

      row_offset += sz + 1

def create(tsv_path, save_path, settings):

  with os.scandir(save_path) as files:
    most_recent_file = sorted(list(files), key=lambda x: x.stat().st_ctime_ns)[0]
    recent_wb = load_workbook(most_recent_file.path, read_only=True)
    recent_ws = recent_wb.active

  old_orders = load_xlsx(recent_ws)
  new_orders = load_tsv(tsv_path)

  orders = transfer_user_notes(old_orders, new_orders)

  wb = Workbook()
  ws = wb.active
  write(ws, orders, TODAY, False, settings)

  wb.save('{}{}'.format(save_path, OUTPUT_FILENAME))

  return OUTPUT_FILENAME

def update(tsv_path, xlsx_path, settings):

  wb = load_workbook(filename=xlsx_path)
  ws = wb[wb.sheetnames[0]]
  ws_new = wb.create_sheet(title="Orders")

  existing_orders = load_xlsx(ws)
  new_orders = load_tsv(tsv_path)

  updated_orders = compare(existing_orders, new_orders)

  created_on = datetime.datetime.strptime(xlsx_path.split("\\")[-1].split(" ")[0], "%Y-%m-%d").date() #need to delete

  write(ws_new, updated_orders, created_on, True, settings)

  wb.remove(ws)
  wb.save('{}'.format(xlsx_path))

  return xlsx_path
